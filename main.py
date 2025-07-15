# import openpyxl
# from collections import defaultdict

# def parse_faculty_schedule(filename):
#     wb = openpyxl.load_workbook(filename)
#     ws = wb.active

#     # Get course names from row 1, columns 2 onwards
#     courses = [cell.value for cell in ws[1][1:] if cell.value]

#     # Dictionary to hold course → list of profs
#     course_to_profs = defaultdict(list)

#     # Dictionary to count how many courses each prof is assigned
#     prof_to_courses = defaultdict(list)

#     print(" Faculty to Courses:")
#     for row in ws.iter_rows(min_row=5, min_col=1, max_col=ws.max_column):
#         prof_name = row[0].value
#         if not prof_name:
#             continue

#         for i, cell in enumerate(row[1:]):
#             if str(cell.value).strip().upper() == 'X':
#                 course = str(courses[i])
#                 course_to_profs[course].append(prof_name)

#     # Assignment phase
#     assigned_courses = defaultdict(list)  # course → [profs]
#     prof_assignment_count = defaultdict(int)

#     for course, eligible_profs in course_to_profs.items():
#         assigned = []
#         for prof in eligible_profs:
#             if prof_assignment_count[prof] < 4 and len(assigned) < 3:
#                 assigned.append(prof)
#                 prof_assignment_count[prof] += 1
#                 prof_to_courses[prof].append(course)

#         assigned_courses[course] = assigned

#         if len(assigned) < 3:
#             print(f"⚠️ Warning: Could only assign {len(assigned)} prof(s) to course {course}")

#     # Output assignments
#     print("\n Course Assignments:")
#     for course in sorted(assigned_courses.keys()):
#         profs = assigned_courses[course]
#         print(f"{course}: {', '.join(profs)}")

#     print("\n Professor Loads:")
#     for prof, course_list in prof_to_courses.items():
#         print(f"{prof}: {', '.join(course_list)} (Total: {len(course_list)})")

# if __name__ == "__main__":
#     parse_faculty_schedule("faculty.xlsx")

from flask import Flask, render_template, request
import openpyxl
from collections import defaultdict

app = Flask(__name__)

# List of all course codes from your HTML
COURSE_CODES = [
    "6600", "6601", "6602", "6604", "6610", "6612", "6603", "6607", "6608", "6609",
    "6620", "6661", "6662", "6695", "6618", "6628", "6658", "6676", "6706", "6709",
    "6747", "6632", "6651", "6677", "6700", "6637", "6638", "6639", "6654", "6697",
    "6700b", "6701", "6735", "6830", "6831"
]

@app.route('/', methods=['GET', 'POST'])
def course_form():
    if request.method == 'POST':
        course_limits = {}
        for course in COURSE_CODES:
            value = request.form.get(f'course{course}')
            try:
                course_limits[course] = int(value)
            except (TypeError, ValueError):
                course_limits[course] = 3  # default fallback

        print("\n--- Running Assignment with User Inputs ---\n")
        parse_faculty_schedule("faculty.xlsx", course_limits)

    return render_template('form.html')

def parse_faculty_schedule(filename, course_limits):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    courses = [cell.value for cell in ws[1][1:] if cell.value]
    course_to_profs = defaultdict(list)
    prof_to_courses = defaultdict(list)

    print("Faculty to Courses:")
    for row in ws.iter_rows(min_row=5, min_col=1, max_col=ws.max_column):
        prof_name = row[0].value
        if not prof_name:
            continue
        for i, cell in enumerate(row[1:]):
            if str(cell.value).strip().upper() == 'X':
                course = str(courses[i])
                course_to_profs[course].append(prof_name)

    assigned_courses = defaultdict(list)
    prof_assignment_count = defaultdict(int)

    for course, eligible_profs in course_to_profs.items():
        assigned = []
        course_limit = course_limits.get(course, 3)  # default to 3 if not specified
        for prof in eligible_profs:
            if prof_assignment_count[prof] < 4 and len(assigned) < course_limit:
                assigned.append(prof)
                prof_assignment_count[prof] += 1
                prof_to_courses[prof].append(course)
        assigned_courses[course] = assigned

        if len(assigned) < course_limits.get(course, 3):
            print(f"⚠️ Warning: Could only assign {len(assigned)} prof(s) to course {course}")

    print("\nCourse Assignments:")
    for course in sorted(assigned_courses.keys()):
        profs = assigned_courses[course]
        print(f"{course}: {', '.join(profs)}")

    print("\nProfessor Loads:")
    for prof, course_list in prof_to_courses.items():
        print(f"{prof}: {', '.join(course_list)} (Total: {len(course_list)})")

if __name__ == "__main__":
    app.run(debug=True)
